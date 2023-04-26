# -*- coding: utf-8 -*-
"""
Created on Wed May 19 22:52:24 2021

@author: Haad-Rathore
"""

import os
from openpyxl import load_workbook
import pandas as pd
from scipy import signal, interpolate, stats
import seaborn as sns
from matplotlib import pyplot as plt
import numpy as np

def ShowFFT(input_signal,Fs, fig_flag = 1):
    N = len(input_signal)
    fnyquist = Fs/2.0

    # Single-sided magnitude spectrum with frequency in Hz
    X_mags = np.absolute(np.fft.fft(input_signal,norm = 'ortho'))
    bin_vals = np.arange(0,N)
    freq_ax_Hz = bin_vals * Fs/N
    N_2 = int(np.ceil(N/2))
    
    if (fig_flag == 1):
        # plotting
        fig, (ax1, ax2) = plt.subplots(1, 2)
        fig.suptitle('Single-sided Magnitude spectrum (Hz)')
        ax1.plot(freq_ax_Hz[0:N_2],X_mags[0:N_2])
        ax2.plot(freq_ax_Hz[0:N_2],20*np.log10(X_mags[0:N_2]))
        # ax1.xlabel('Frequency (Hz)')
        ax1.set_ylabel('Magnitude')
        ax1.set_xlabel('Frequency (Hz)')
        ax2.set_xlabel('Frequency (Hz)')
        ax2.set_ylabel('Magnitude (dB)')
    
    f = freq_ax_Hz[0:N_2]
    FFT = X_mags[0:N_2]
    
    return f, FFT
    
# 10-200 Hz
def filterSignal_LFP(input_signal,Fs):
    # Prep
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    if len(input_signal.shape) != 1:
        axis_value = 1
    else:
        axis_value = 0
    # Loading filter data
    filename_FilterData = os.path.join(os.getcwd(),'filterData.xlsx')
    df = pd.read_excel(filename_FilterData,sheet_name = 5, header = None)
    X = df.to_numpy()
    SoS = X[0:12,0:6]               # Hard coded SOS matrix imported from Excel file
    ScaleValues = X[:,7]            # Hard coded G scale values imported from Excel file
    SoS = SoS.copy(order = 'C')     # because python imports this matrix as Fortran order style
    
    signal_out = signal.sosfiltfilt(SoS,input_signal, axis = axis_value) * np.prod(ScaleValues)
    return signal_out    

# ----------------- Good filters ------------------------------#
# Notch filter 60
def filterSignal_notch(input_signal, Fs, C0 = 60, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    Q = 20                     # Quality factor determines bandwidth
    b,a = signal.iirnotch(C0,Q, fs = Fs)  # IIR comb filter
    signal_out = signal.filtfilt(b,a, input_signal, axis = axis_value)
    return signal_out

# Low pass <2 Hz
def filterSignal_lowpass(input_signal, Fs, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    cutoff_low = 1.5                  # Low pass freq for LFP band
    sos = signal.butter(5, cutoff_low, btype = 'lowpass', output = 'sos', fs = Fs)  # IIR filter
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

def filter_Savitzky_fast(input_signal):
    signal_out = signal.savgol_filter(input_signal,9,3)
    return signal_out

def filter_Savitzky_slow(input_signal):
    signal_out = signal.savgol_filter(input_signal,19,3)
    return signal_out

# 13 - 160 Hz
def filterSignal_lowpassLFP(input_signal, Fs, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    cutoff_high = 160                # High pass freq for LFP band
    cutoff_low = 13                  # Low pass freq for LFP band
    # cutoff = cutoff / (0.5*Fs)
    sos = signal.butter(20, [cutoff_low,cutoff_high], btype = 'bandpass', output = 'sos', fs = Fs)  # IIR filter
    
    
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

# 8 - 80 Hz
def filterSignal_BP_LFP(input_signal, Fs, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    cutoff_high = 80                # High pass freq for LFP band
    cutoff_low = 8                  # Low pass freq for LFP band
    # cutoff = cutoff / (0.5*Fs)
    sos = signal.butter(10, [cutoff_low,cutoff_high], btype = 'bandpass', output = 'sos', fs = Fs)  # IIR filter
    
    
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

# 300 to 3 KHz
def filterSignal_MUA(input_signal,Fs, axis_value = 0):
    # Prep
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    # Loading filter data
    # filename_FilterData = os.path.join(os.getcwd(),'filterData.xlsx')
    # df = pd.read_excel(filename_FilterData,sheet_name = 4, header = None)
    # X = df.to_numpy()
    # SoS = X[0:197,0:6]                          # Hard coded SOS matrix imported from Excel file
    # ScaleValues = X[:,7]                        # Hard coded G scale values imported from Excel file
    # SoS = SoS.copy(order = 'C')                 # because python imports this matrix as Fortran order style
    # b,a = signal.sos2tf(SoS)                    # Convert to b, a coeeficients
    # b = b * np.prod(ScaleValues)                # Scaling 
    # signal_out = signal.filtfilt(b,a, input_signal, axis = axis_value)
    # signal_out = signal.sosfiltfilt(SoS,input_signal, axis = axis_value) * np.prod(ScaleValues)
    cutoff = np.array([300,3000])
    sos = signal.butter(20, cutoff, btype = 'bandpass', output = 'sos', fs = Fs)  # IIR filter
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

# High-Pass at 12 Hz
def filterSignal_High(input_signal, Fs, cutoff, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    sos = signal.butter(6, cutoff, btype = 'highpass', output = 'sos', fs = Fs)  # IIR filter
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

def compute_PSD(arr, Fs, window_length, axis_value = 0):
    # arr is the input data array
    # Rows index is trial # ie different data sets 
    # Column index denotes time (in seconds)
    # Fs : sampling frequency in Hz
    # window_length : Length of the window for chopping the sequence into pieces 
    # axis : the axis along which the spectrogram will be computed. (the time axis)
    truncate_window = signal.windows.kaiser(window_length,7.6,sym = False)
    # truncate_window = signal.windows.kaiser(window_length,8.6,sym = False)
    Noverlap = np.ceil(window_length/2)
    f, t, Sxx = signal.spectrogram(arr, fs=Fs, window = truncate_window, noverlap = Noverlap, scaling = 'density', axis = axis_value, mode = 'psd')
    return f, t, Sxx
    
def detect_peak_basic(arr, num, x):
    # arr is the array of values from which the peak will be detected
    # num is the number of peaks to consider when detecting the main peak
    # x is the threshold for peak detection
    # Basic method of finding the first highest peak and their indices
    arr_max_indx = np.argpartition(arr,-num)[-num:]
    arr_max_indx = arr_max_indx[np.argsort(arr[arr_max_indx])]
    
    
    arr_max = arr[arr_max_indx[:]]
    arr_max = arr_max[arr_max > x]
    
    if arr_max.size == 0:
        return x, 2000
    else:
        arr_max_indx = np.flip(arr_max_indx[np.arange(-1,-arr_max.size-1,-1)])
        peak_val = arr[arr_max_indx[-1]]
        for iter in np.flip(range(arr_max.size)):
            pd = (peak_val - arr[arr_max_indx[iter]])/peak_val
            if pd < 0.08:
                pk_indx_new = iter
        
        arr_max_indx = arr_max_indx[pk_indx_new]
        arr_max = arr_max[pk_indx_new]
        return arr_max, arr_max_indx
    
    
    # if (arr_max_indx[-1] < arr_max_indx[-2]):
    #     arr_max_indx = arr_max_indx[-1]
    #     arr_max = arr[arr_max_indx] 
    # else:
    #     pd = (arr[arr_max_indx[-1]] - arr[arr_max_indx[-2]])/arr[arr_max_indx[-2]]
    #     if pd < 0.15:
    #         arr_max_indx = arr_max_indx[-2]
    #         arr_max = arr[arr_max_indx]
    #     else:
    #         arr_max_indx = arr_max_indx[-1]
    #         arr_max = arr[arr_max_indx]
    

def interp_chan_loss(data_in, shank_missed):
    """
    NTERP_CHAN_LOSS Interpolate voltage values of missing channels in a
    linear array of electrodes
       The function takes as its input a 2D array with rows as time and
       columns as electrodes. Missing electrodes would therefore be signified
       by a column of NaNs.
       
       INPUT data_in : 2D matrix containing time and electrode information as rows
       and columns, respectively. Each column is therefore a new electrode.
       INPUT shank_missed : An array denoting what channels were missing. Filled
       with NaNs
    """
    
    # Actual output axis
    x = np.arange(0,data_in.shape[1])

    # Input y with missing values
    indx_not_missing = ~np.isnan(shank_missed)
    data_in = data_in[:,indx_not_missing]
    # Input x with missing values
    shank_missed = shank_missed[~np.isnan(shank_missed)]
    # Interpolation along the 1st axis (along cortical depth)
    f_data_out = interpolate.interp1d(shank_missed,data_in,kind = 'quadratic',axis = 1, fill_value='extrapolate')
    return f_data_out(x)

def CSD_compute(data, SR, spacing, conductivity = 0.3):
    """
    # Function inputs (required)
    
        1) data: input data where columns contain temporal data and rows
          contain spatial data (i.e each column is a voltage trace from a
          single electrode). The data must be in volts.
        2) SR: sampling rate of the input data, in Hz.
        3) spacing: this is the spacing between two adjacent electrodes.     
          Must be in meters. 
    
    % Function inputs (optional: name-value pair arguments)
    
       1) 'conductivity': the conductivity of the extracellular medium, in 
          siemans per meter. Default is 0.3 S.m^-1. 
       2) 'inverse': obtains the CSD using the inverse CSD method. This
          option requires the radius (surrounding each electrode 
          contact) in which the CSD is considered to be restricted to. 
          Frequently taken as the radius of the electrode or a multiple
          (e.g. 5) of the electrode spacing. 
          
     Function outputs (CSDoutput)
        
         The output of the function consists of the current source density in
         amps.meters^-3. Default units are microamps.millimeters^-3. The CSD
         data is in the same format as the input voltage data (temporal data 
         in columns; spatial data in rows). When using the standard CSD
         method, the first and last columns (consisting of data from the 1st 
         and last electrodes) will be filled with NaNs. This is due to the
         standard CSD method unable to obtain the CSD at these outermost 
         electrode contacts. The inverse CSD method does not have this 
         limitation.
    """
    conA=1000000
    conL=1000
    SPm = (1/SR)*1000   # sampling period in ms      
    xAxis = np.arange(SPm,SPm*(len(data)+1),SPm)  # x-axis for plots
    
    radius = 10*spacing # the radius (surrounding each electrode contact) in which the CSD is considered to be restricted to
    
    # Obtain CSD
    numElec = data.shape[1]
    z = np.arange(spacing,spacing*(numElec+1),spacing)
    F = np.zeros((numElec,numElec))
    for ii in range(0,numElec):           # generating the F matrix
        for ij in range(0,numElec):
            F[ii,ij] = (spacing/(2*conductivity)) * (np.sqrt(np.square(z[ij]-z[ii])+np.square(radius))-np.abs(z[ij]-z[ii]))
    CSD_a = np.matmul(np.linalg.inv(F),np.transpose(data))
    CSD_a = CSD_a / (np.power(conL,3)) # converts CSD units to desired length (m, mm, etc)
    CSD_a = CSD_a * conA # converts CSD units to desired amps (A, mA, uA, etc)
    # CSD_a = np.transpose(CSD_a) 
    return CSD_a      

# Read from whisker_stim.txt file (generated by my Labview program)
def read_stimtxt(matlabTXT):
    # Read .txt file
    Txt_File = open(matlabTXT,"r")
    str1 = Txt_File.readlines()
    stim_start_time = float(str1[3]) - 0.05       # Stimulation start time (50 ms error in labview)
    stim_num = int(str1[6])                 # Number of stimulations
    seq_period = float(str1[4])             # Period of each sequence consisting of 10 to 15 frames
    len_trials = int(str1[1])               # Length of a single trial in # seq
    num_trials = int(str1[7])               # total number of trials
    FramePerSeq = int(str1[2])              # frames per sequence. Important info to read the data.timing file
    total_seq = num_trials * len_trials
    len_trials_arr = list(range(1,total_seq,len_trials))
    print('Each sequence lasts...',seq_period,'sec')
    print('Number of sequences in each trial...',len_trials)
    print('Total number of trials in this dataset are:', num_trials)
    print('Reading the total number of sequences in this dataset...',total_seq,'\n.\n.\n.')
    Txt_File.close()
    
    return stim_start_time, stim_num, seq_period, len_trials, num_trials, FramePerSeq, total_seq, len_trials_arr

def toggle_plot(fig):
  # This function is called by a keypress to hide/show the figure
  fig.set_visible(not fig.get_visible())
  plt.draw()
  
def zscore_bsl(time_series,bsl_mean,bsl_std):
    time_series_out = (time_series - bsl_mean)/bsl_std
    return time_series_out

def bsl_norm(time_series):
    bsl_mean = np.mean(time_series[0:3])
    if bsl_mean == 0:
        time_series_out = np.nan * np.ones(time_series.shape) 
    else:
        time_series_out = (time_series - bsl_mean)/bsl_mean
        
    return time_series_out

def plot_all_trials(input_arr,Fs,folder_path,clus_dict):
    # INPUT input_arr is a 1D array with avg FR of a single cluster
    # INPUT Fs is the sampling frequency representing the time axis
    # INPUT folder_path is the output folder where the figures will be saved
    # INPUT clus_dict contains the information on the cluster 
    
    # input_arr = stats.zscore(input_arr)
    t_axis = np.linspace(0,input_arr.shape[0]/Fs,input_arr.shape[0])
    t_start_indx = np.squeeze(np.where(t_axis >= 1.8))[0]      # stim start set to 2.45 seconds
    t_end_indx = np.squeeze(np.where(t_axis <= 6.5))[-1]        # stim end set to 5.15 seconds
    filename_save = os.path.join(folder_path,'FR_' + str(clus_dict['cluster_id']) + '.png')

    
    f, a = plt.subplots(1,1)
    a.set_ylabel('FR/Hz')
    len_str = 'Cluster ID:' + str(clus_dict['cluster_id']) + '| Shank:' + str(clus_dict['shank_num']) + '| Depth:' + str(clus_dict['prim_ch_coord'][1])
    f.suptitle(len_str)
    if clus_dict['clus_prop'] == 1:
        a.plot(t_axis[t_start_indx:t_end_indx+25],input_arr[t_start_indx:t_end_indx+25],'g', lw=2.0)
    elif (clus_dict['clus_prop'] == -1):
        a.plot(t_axis[t_start_indx:t_end_indx+25],input_arr[t_start_indx:t_end_indx+25],'b', lw=2.0)
    else:
        a.plot(t_axis[t_start_indx:t_end_indx+25],input_arr[t_start_indx:t_end_indx+25],'k', lw=2.0)

    plt.axvline(2.2,linestyle = 'dashed', linewidth = 2.1)
    plt.axvline(5,linestyle = 'dashed', linewidth = 2.1)
    # a.set_yticks([])
    f.set_size_inches((5, 3), forward=False)
    plt.savefig(filename_save,format = 'png')
    plt.close(f)
    
    return None

# Write to excel files Append mode perfected
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl',mode='a',if_sheet_exists = 'overlay' )

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow,**to_excel_kwargs)

    # save the workbook
    writer.save()
    
# The following is Jiaao's code for determining possible candidates for merging
# TRANSIENT_AMPLITUDE_VALID_DURATION = 10e-4 # seconds (duration of data before and after each spike that we consider when deciding the transient amplitude)
# tavd_nsample = int(np.ceil(TRANSIENT_AMPLITUDE_VALID_DURATION*F_SAMPLE))
MAX_GEOM_DIST = 25 # um
ADJACENCY_RADIUS_SQUARED = 140**2
def get_peak_amp_ratio_matrix(data_a, data_b=None):
    """
    assumes data_a is (n_1, ) and data_b is (n_2, )
    returns (n_1, n_2) distance matrix
    where n_1 and n_2 could be cluster counts from 2 sessions
    """
    if data_b is None:
        data_b = data_a
    data_min = np.minimum(data_a[:,None], data_b[None,:])
    data_max = np.maximum(data_a[:,None], data_b[None,:])
    return data_max/data_min - 1.

# The following is Jiaao's code for determining possible candidates for merging
def calc_key_metrics(templates_full, firings, geom, f_sample, radius_squared=ADJACENCY_RADIUS_SQUARED):
    """Calculate key cluster metrics used by recursive automerging"""
    # Must input curated templates.mda numpy array and firings.mda numpy array 
    
    TRANSIENT_AMPLITUDE_VALID_DURATION = 10e-4 # seconds (duration of data before and after each spike that we consider when deciding the transient amplitude)
    tavd_nsample = int(np.ceil(TRANSIENT_AMPLITUDE_VALID_DURATION*f_sample))
    n_chs, waveform_len, n_clus = templates_full.shape
    
    # get primary channels
    pri_ch_lut = -1 * np.ones(n_clus, dtype=int)
    n_pri_ch_known = 0
    for (spk_ch, spk_lbl) in zip(firings[0,:], firings[2,:]):
        spk_lbl = int(spk_lbl)
        if pri_ch_lut[spk_lbl-1]==-1:
            pri_ch_lut[spk_lbl-1] = spk_ch-1
            n_pri_ch_known += 1
            if n_pri_ch_known==n_clus:
                break
    
    # slice templates
    my_slice = slice(int(waveform_len//2-tavd_nsample), int(waveform_len//2+tavd_nsample), 1)
    templates = templates_full[:,my_slice,:]
    # waveform_len_sliced = templates.shape[1]

    # get template peaks and p2ps
    template_peaks = np.max(np.abs(templates), axis=1)
    print("template_peaks shape <should be (n_ch,n_clus)>:", template_peaks.shape)
    peak_amplitudes = template_peaks[pri_ch_lut, np.arange(n_clus)] # (n_clus,)
    template_peaks = np.max(templates, axis=1) # use full to calculate p2p
    template_troughs = np.min(templates, axis=1)
    template_p2ps = template_peaks - template_troughs

    # estimate locations by center-of-mass
    clus_coordinates = np.zeros((n_clus, 2))
    for i_clus in range(n_clus):
        prim_ch = pri_ch_lut[i_clus]
        prim_x, prim_y = geom[prim_ch, :]
        non_neighbor_mask = ((geom[:,0]-prim_x)**2 + (geom[:,1]-prim_y)**2 >= radius_squared)
        weights = template_p2ps[:, i_clus]
        weights[non_neighbor_mask] = 0
        weights = weights / np.sum(weights)
        clus_coordinates[i_clus, :] = np.sum(weights[:,None] * geom, axis=0)
    
    return templates, pri_ch_lut, peak_amplitudes, clus_coordinates

# The following is Jiaao's code for determining possible candidates for merging
def calc_merge_candidates(templates, locations, peak_amplitudes):
    """
    calculate merging candidates by distance & waveform similarity\n
    returns a 3-column matrix of (n_pairs), the columns would be (src_unit, snk_unit, cand?)\n
    Please Use sliced templates and clean firings only (noise cluster must be rejected and the clean ones reordered)
    """
    # Must input curated templates.mda numpy array 

    n_ch, waveform_len, n_clus = templates.shape
    template_features = templates.reshape((n_ch*waveform_len, n_clus)).T

    pairs_all = []
    pairs_cand = []
    for i_clus in range(n_clus):
        neighborhood_mask = np.sum((locations-locations[i_clus,:])**2, axis=1) < MAX_GEOM_DIST**2
        neighborhood_mask[i_clus: ] = False # Force non-directed graph for merging; also no comparison with self
        n_neighborhood = np.sum(neighborhood_mask)
        if n_neighborhood<1:
            continue
        neighborhood_clus_ids = np.where(neighborhood_mask)[0] + 1 # cluster id starts from 1
        current_clus_id = i_clus + 1        # Cluster IDs start from 1
        dist_mat = np.array([np.corrcoef(template_features[i_clus,:], template_features[k-1, :])[1,0] for k in neighborhood_clus_ids])
        corr_mask = dist_mat > 0.7 # actually a vector
        amp_ratio_mat = get_peak_amp_ratio_matrix(peak_amplitudes[:,None][i_clus,:], peak_amplitudes[neighborhood_mask]).squeeze()
        amp_ratio_mask = amp_ratio_mat < 0.5 # np.logical_and(amp_ratio_mat>0.8, amp_ratio_mat<1.25)
        merge_cand_mask = np.logical_and(amp_ratio_mask, corr_mask) # actually a vector
        n_cands = np.sum(merge_cand_mask)
        
        clus_id_paired_prev_all, clus_id_paired_post_all = np.zeros(n_neighborhood, dtype=int)+current_clus_id, neighborhood_clus_ids
        clus_id_paired_prev_cand, clus_id_paired_post_cand = np.zeros(n_cands, dtype=int)+current_clus_id, neighborhood_clus_ids[merge_cand_mask]

        pairs_all.extend(list(zip(clus_id_paired_prev_all, clus_id_paired_post_all))) # list of tuples
        pairs_cand.extend(list(zip(clus_id_paired_prev_cand, clus_id_paired_post_cand))) # list of tuples

        # plt.figure(figsize=(12,4)); 
        # plt.subplot(131); plt.imshow(merge_cand_mask, cmap='gray'); plt.colorbar(); 
        # plt.xticks(np.arange(n_neighborhood), neighborhood_clus_ids+1)
        # plt.yticks(np.arange(n_neighborhood), neighborhood_clus_ids+1)
        # plt.subplot(132); plt.imshow(dist_mat, cmap='gray', vmin=0, vmax=1); plt.colorbar(); plt.title("Corr")
        # plt.xticks(np.arange(n_neighborhood), neighborhood_clus_ids+1)
        # plt.yticks(np.arange(n_neighborhood), neighborhood_clus_ids+1)
        # plt.subplot(133); plt.imshow(amp_ratio_mat, cmap='gray'); plt.colorbar(); plt.title("ampRatio")
        # plt.xticks(np.arange(n_neighborhood), neighborhood_clus_ids+1)
        # plt.yticks(np.arange(n_neighborhood), neighborhood_clus_ids+1)
        # plt.show()
    if len(pairs_all) > 0: 
        cand_mask_1d = np.array([(pair in pairs_cand) for pair in pairs_all], dtype=bool) # (n_pairs,)
        assert(np.sum(cand_mask_1d)==len(pairs_cand))
        arr_pairs_all = np.array(pairs_all) # (n_pairs,2)
        print(arr_pairs_all.shape,cand_mask_1d.shape)
        arr_ret = np.concatenate([arr_pairs_all, cand_mask_1d[:,None]], axis=1)
    else:
        arr_ret = None
    return arr_ret
    
def makeSymmetric(mat):
    assert mat.shape[0] == mat.shape[1], \
        print('Matrix is not a square matrix')
    N = mat.shape[0]
    # Loop to traverse lower triangular
    # elements of the given matrix
    for i in range(0, N):
        for j in range(0, N):
            if (j < i):
                mat[i][j] = mat[j][i] = (mat[i][j] +
                                         mat[j][i]) / 2
                
    return mat