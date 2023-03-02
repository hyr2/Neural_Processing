# -*- coding: utf-8 -*-
"""
Created on Wed May 19 22:52:24 2021

@author: Haad-Rathore
"""

import os
from openpyxl import load_workbook
import pandas as pd
from scipy import signal, interpolate

from matplotlib import pyplot as plt
import numpy as np

def ShowFFT(input_signal,Fs, fig_flag = 1):
    N = len(input_signal)
    fnyquist = Fs/2.0

    # Single-sided magnitude spectrum with frequency in Hz
    X_mags = np.absolute(np.fft.fft(input_signal,norm = 'ortho'))
    bin_vals = np.arange(0,N)
    freq_ax_Hz = bin_vals * Fs/N
    N_2 = int(np.ceil(N/2))
    
    if (fig_flag == 1):
        # plotting
        fig, (ax1, ax2) = plt.subplots(1, 2)
        fig.suptitle('Single-sided Magnitude spectrum (Hz)')
        ax1.plot(freq_ax_Hz[0:N_2],X_mags[0:N_2])
        ax2.plot(freq_ax_Hz[0:N_2],20*np.log10(X_mags[0:N_2]))
        # ax1.xlabel('Frequency (Hz)')
        ax1.set_ylabel('Magnitude')
        ax1.set_xlabel('Frequency (Hz)')
        ax2.set_xlabel('Frequency (Hz)')
        ax2.set_ylabel('Magnitude (dB)')
    
    f = freq_ax_Hz[0:N_2]
    FFT = X_mags[0:N_2]
    
    return f, FFT
    
# 10-200 Hz
def filterSignal_LFP(input_signal,Fs):
    # Prep
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    if len(input_signal.shape) != 1:
        axis_value = 1
    else:
        axis_value = 0
    # Loading filter data
    filename_FilterData = os.path.join(os.getcwd(),'filterData.xlsx')
    df = pd.read_excel(filename_FilterData,sheet_name = 5, header = None)
    X = df.to_numpy()
    SoS = X[0:12,0:6]               # Hard coded SOS matrix imported from Excel file
    ScaleValues = X[:,7]            # Hard coded G scale values imported from Excel file
    SoS = SoS.copy(order = 'C')     # because python imports this matrix as Fortran order style
    
    signal_out = signal.sosfiltfilt(SoS,input_signal, axis = axis_value) * np.prod(ScaleValues)
    return signal_out    

# ----------------- Good filters ------------------------------#
# Notch filter 60
def filterSignal_notch(input_signal, Fs, C0 = 60, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    Q = 20                     # Quality factor determines bandwidth
    b,a = signal.iirnotch(C0,Q, fs = Fs)  # IIR comb filter
    signal_out = signal.filtfilt(b,a, input_signal, axis = axis_value)
    return signal_out

# 13 - 160 Hz
def filterSignal_lowpassLFP(input_signal, Fs, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    cutoff_high = 160                # High pass freq for LFP band
    cutoff_low = 13                  # Low pass freq for LFP band
    # cutoff = cutoff / (0.5*Fs)
    sos = signal.butter(20, [cutoff_low,cutoff_high], btype = 'bandpass', output = 'sos', fs = Fs)  # IIR filter
    
    
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

# 8 - 80 Hz
def filterSignal_BP_LFP(input_signal, Fs, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    cutoff_high = 80                # High pass freq for LFP band
    cutoff_low = 8                  # Low pass freq for LFP band
    # cutoff = cutoff / (0.5*Fs)
    sos = signal.butter(10, [cutoff_low,cutoff_high], btype = 'bandpass', output = 'sos', fs = Fs)  # IIR filter
    
    
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

# 300 to 3 KHz
def filterSignal_MUA(input_signal,Fs, axis_value = 0):
    # Prep
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    # if len(input_signal.shape) != 1:
    #     axis_value = 1
    # else:
    #     axis_value = 0
    # Loading filter data
    # filename_FilterData = os.path.join(os.getcwd(),'filterData.xlsx')
    # df = pd.read_excel(filename_FilterData,sheet_name = 4, header = None)
    # X = df.to_numpy()
    # SoS = X[0:197,0:6]                          # Hard coded SOS matrix imported from Excel file
    # ScaleValues = X[:,7]                        # Hard coded G scale values imported from Excel file
    # SoS = SoS.copy(order = 'C')                 # because python imports this matrix as Fortran order style
    # b,a = signal.sos2tf(SoS)                    # Convert to b, a coeeficients
    # b = b * np.prod(ScaleValues)                # Scaling 
    # signal_out = signal.filtfilt(b,a, input_signal, axis = axis_value)
    # signal_out = signal.sosfiltfilt(SoS,input_signal, axis = axis_value) * np.prod(ScaleValues)
    cutoff = np.array([300,3000])
    sos = signal.butter(20, cutoff, btype = 'bandpass', output = 'sos', fs = Fs)  # IIR filter
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

# High-Pass at 12 Hz
def filterSignal_High(input_signal, Fs, cutoff, axis_value = 0):
    signal_out = np.empty((input_signal.shape),dtype=np.single)
    sos = signal.butter(6, cutoff, btype = 'highpass', output = 'sos', fs = Fs)  # IIR filter
    signal_out = signal.sosfiltfilt(sos, input_signal, axis = axis_value)
    return signal_out

def compute_PSD(arr, Fs, window_length, axis_value = 0):
    # arr is the input data array
    # Rows index is trial # ie different data sets 
    # Column index denotes time (in seconds)
    # Fs : sampling frequency in Hz
    # window_length : Length of the window for chopping the sequence into pieces 
    # axis : the axis along which the spectrogram will be computed. (the time axis)
    truncate_window = signal.windows.kaiser(window_length,7.6,sym = False)
    # truncate_window = signal.windows.kaiser(window_length,8.6,sym = False)
    Noverlap = np.ceil(window_length/2)
    f, t, Sxx = signal.spectrogram(arr, fs=Fs, window = truncate_window, noverlap = Noverlap, scaling = 'density', axis = axis_value, mode = 'psd')
    return f, t, Sxx
    
def detect_peak_basic(arr, num, x):
    # arr is the array of values from which the peak will be detected
    # num is the number of peaks to consider when detecting the main peak
    # x is the threshold for peak detection
    # Basic method of finding the first highest peak and their indices
    arr_max_indx = np.argpartition(arr,-num)[-num:]
    arr_max_indx = arr_max_indx[np.argsort(arr[arr_max_indx])]
    
    
    arr_max = arr[arr_max_indx[:]]
    arr_max = arr_max[arr_max > x]
    
    if arr_max.size == 0:
        return x, 2000
    else:
        arr_max_indx = np.flip(arr_max_indx[np.arange(-1,-arr_max.size-1,-1)])
        peak_val = arr[arr_max_indx[-1]]
        for iter in np.flip(range(arr_max.size)):
            pd = (peak_val - arr[arr_max_indx[iter]])/peak_val
            if pd < 0.08:
                pk_indx_new = iter
        
        arr_max_indx = arr_max_indx[pk_indx_new]
        arr_max = arr_max[pk_indx_new]
        return arr_max, arr_max_indx
    
    
    # if (arr_max_indx[-1] < arr_max_indx[-2]):
    #     arr_max_indx = arr_max_indx[-1]
    #     arr_max = arr[arr_max_indx] 
    # else:
    #     pd = (arr[arr_max_indx[-1]] - arr[arr_max_indx[-2]])/arr[arr_max_indx[-2]]
    #     if pd < 0.15:
    #         arr_max_indx = arr_max_indx[-2]
    #         arr_max = arr[arr_max_indx]
    #     else:
    #         arr_max_indx = arr_max_indx[-1]
    #         arr_max = arr[arr_max_indx]
  

def interp_session_loss(data_in, day_local_axis, day_axis_ideal):
    """
    NTERP_SESSION_LOSS Interpolate data values of missing sessions in a
    longitudinal stroke study
       The function takes as its input a 2D array with rows as sessions and
       columns as shanks. The function does nearest neighbour interpolation 
       for days less than 28. For days >= 28, averaging of the existing data 
       points is performed to fill the missing data points.
       
       INPUT data_in : 2D matrix containing time and shank information as rows
       and columns, respectively. Each column is therefore a new shank.
       INPUT day_local_axis : An array for the days the recordings took place
       INPUT day_axis_ideal: An array for the ideal days (used for interpolation)
    """
    
    # 2nd dimension is extracted here: 4 for shanks and 3 for cell types
    YY_dim = data_in.shape[1]
    
    # day_local_axis changed to include all three baselines
    if not any(np.isin(day_local_axis,-3)):
        day_local_axis = np.insert(day_local_axis,0,-3)
        
    local_indx = np.squeeze(np.where(day_local_axis < 28 ))
    # if before day 28 (performing nearest neighbour)    
    f_data_out = interpolate.interp1d(day_local_axis[local_indx],data_in[local_indx,:],kind = 'nearest',axis = 0, fill_value='extrapolate')
    local_indx = np.squeeze(np.where(day_axis_ideal < 28 ))
    out_arr = np.rint(f_data_out(day_axis_ideal[local_indx]))
    
    # if after day 28 (performing average value of post day 28)
    local_indx = np.squeeze(np.where(day_local_axis >= 28 ))
    avg_local = np.rint(np.nanmean(np.reshape(data_in[local_indx,:],[local_indx.size,YY_dim]),axis = 0))
    tmp_indx = np.isin(day_axis_ideal,day_local_axis[local_indx])
    tmp_indx = ~tmp_indx[7:]    # day 28 is a hard coded (requested by Dr.Lan)
    out_arr_28 = np.zeros([tmp_indx.size,YY_dim])
    out_arr_28[tmp_indx,:] = avg_local
    
    # merging two
    merged_arr = np.vstack((out_arr,out_arr_28))
    local_indx = ~np.isin(day_axis_ideal,day_local_axis)
    local_indx[:7] = True
    local_indx = ~local_indx
    
    local_indx_tmp = np.squeeze(np.where(day_local_axis >= 28 ))
    data_fill = data_in[local_indx_tmp,:]
    merged_arr[local_indx,:] = data_fill
    
    out_arr = merged_arr
    # out_arr = np.rint(f_data_out(day_axis_ideal))
    return out_arr


def interp_chan_loss(data_in, shank_missed):
    """
    NTERP_CHAN_LOSS Interpolate voltage values of missing channels in a
    linear array of electrodes
       The function takes as its input a 2D array with rows as time and
       columns as electrodes. Missing electrodes would therefore be signified
       by a column of NaNs.
       
       INPUT data_in : 2D matrix containing time and electrode information as rows
       and columns, respectively. Each column is therefore a new electrode.
       INPUT shank_missed : An array denoting what channels were missing. Filled
       with NaNs
    """
    
    # Actual output axis
    x = np.arange(0,data_in.shape[1])

    # Input y with missing values
    indx_not_missing = ~np.isnan(shank_missed)
    data_in = data_in[:,indx_not_missing]
    # Input x with missing values
    shank_missed = shank_missed[~np.isnan(shank_missed)]
    # Interpolation along the 1st axis (along cortical depth)
    f_data_out = interpolate.interp1d(shank_missed,data_in,kind = 'quadratic',axis = 1, fill_value='extrapolate')
    return f_data_out(x)

def CSD_compute(data, SR, spacing, conductivity = 0.3):
    """
    # Function inputs (required)
    
        1) data: input data where columns contain temporal data and rows
          contain spatial data (i.e each column is a voltage trace from a
          single electrode). The data must be in volts.
        2) SR: sampling rate of the input data, in Hz.
        3) spacing: this is the spacing between two adjacent electrodes.     
          Must be in meters. 
    
    % Function inputs (optional: name-value pair arguments)
    
       1) 'conductivity': the conductivity of the extracellular medium, in 
          siemans per meter. Default is 0.3 S.m^-1. 
       2) 'inverse': obtains the CSD using the inverse CSD method. This
          option requires the radius (surrounding each electrode 
          contact) in which the CSD is considered to be restricted to. 
          Frequently taken as the radius of the electrode or a multiple
          (e.g. 5) of the electrode spacing. 
          
     Function outputs (CSDoutput)
        
         The output of the function consists of the current source density in
         amps.meters^-3. Default units are microamps.millimeters^-3. The CSD
         data is in the same format as the input voltage data (temporal data 
         in columns; spatial data in rows). When using the standard CSD
         method, the first and last columns (consisting of data from the 1st 
         and last electrodes) will be filled with NaNs. This is due to the
         standard CSD method unable to obtain the CSD at these outermost 
         electrode contacts. The inverse CSD method does not have this 
         limitation.
    """
    conA=1000000
    conL=1000
    SPm = (1/SR)*1000   # sampling period in ms      
    xAxis = np.arange(SPm,SPm*(len(data)+1),SPm)  # x-axis for plots
    
    radius = 10*spacing # the radius (surrounding each electrode contact) in which the CSD is considered to be restricted to
    
    # Obtain CSD
    numElec = data.shape[1]
    z = np.arange(spacing,spacing*(numElec+1),spacing)
    F = np.zeros((numElec,numElec))
    for ii in range(0,numElec):           # generating the F matrix
        for ij in range(0,numElec):
            F[ii,ij] = (spacing/(2*conductivity)) * (np.sqrt(np.square(z[ij]-z[ii])+np.square(radius))-np.abs(z[ij]-z[ii]))
    CSD_a = np.matmul(np.linalg.inv(F),np.transpose(data))
    CSD_a = CSD_a / (np.power(conL,3)) # converts CSD units to desired length (m, mm, etc)
    CSD_a = CSD_a * conA # converts CSD units to desired amps (A, mA, uA, etc)
    # CSD_a = np.transpose(CSD_a) 
    return CSD_a      

# Read from whisker_stim.txt file (generated by my Labview program)
def read_stimtxt(matlabTXT):
    # Read .txt file
    Txt_File = open(matlabTXT,"r")
    str1 = Txt_File.readlines()
    stim_start_time = float(str1[3]) - 0.05       # Stimulation start time (50 ms error in labview)
    stim_num = int(str1[6])                 # Number of stimulations
    seq_period = float(str1[4])             # Period of each sequence consisting of 10 to 15 frames
    len_trials = int(str1[1])               # Length of a single trial in # seq
    num_trials = int(str1[7])               # total number of trials
    FramePerSeq = int(str1[2])              # frames per sequence. Important info to read the data.timing file
    total_seq = num_trials * len_trials
    len_trials_arr = list(range(1,total_seq,len_trials))
    print('Each sequence lasts...',seq_period,'sec')
    print('Number of sequences in each trial...',len_trials)
    print('Total number of trials in this dataset are:', num_trials)
    print('Reading the total number of sequences in this dataset...',total_seq,'\n.\n.\n.')
    Txt_File.close()
    
    return stim_start_time, stim_num, seq_period, len_trials, num_trials, FramePerSeq, total_seq, len_trials_arr

# Write to excel files Append mode perfected
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl',mode='a',if_sheet_exists='overlay')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow,**to_excel_kwargs)

    # save the workbook
    writer.save()
    
def bsl_norm(time_series):
    bsl_mean = np.mean(time_series[0:3])
    if bsl_mean == 0:
        time_series_out = np.nan * np.ones(time_series.shape) 
    else:
        time_series_out = (time_series - bsl_mean)/bsl_mean
        
    return time_series_out